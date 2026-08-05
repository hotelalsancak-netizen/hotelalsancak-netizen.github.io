#!/usr/bin/env python3
"""
kart_yukle.py — bir haftalık kapı-kilidi ZIP'ini alıp "Satılmadan kullanılan odalar"
raporu olarak yayınlar. Resepsiyon PC'sindeki program el terminalindeki verileri
okuyup PDF'leri klasöre kaydedip zipliyor (örn. "20.07.2026 kapı-...zip"); bu program
gerisini yapar:

  1. ZIP'i (veya klasörü) açar, içindeki oda PDF'lerini bulur.
  2. Kart okumalarını çözer (parse_cards / pdftotext).
  3. Haftayı otomatik belirler (kart tarihlerinin ISO haftası; --week ile değiştirilebilir).
  4. O haftanın oda planını (doluluk) Elektra'dan çeker.
  5. Oda değişimi verisini (varsa) kullanır — yoksa rapora uyarı düşer.
  6. Raporu üretir, DASH_PASSWORD ile ŞİFRELER, o haftanın bloğu olarak kaydeder.
  7. Hafta listesini (index) günceller; commit + push + bulut yayınını tetikler.

Kullanım:
    python3 kart_yukle.py "~/Downloads/20.07.2026 kapı-....zip"
    python3 kart_yukle.py cardreads/13072026            # klasör de olur
    python3 kart_yukle.py <zip> --week 2026-07-13        # haftayı elle ver
    python3 kart_yukle.py <zip> --no-publish             # sadece yerelde üret
"""
import argparse
import datetime as dt
import glob
import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

import dashcrypto
import elektra_api as E
import parse_cards
import build_report
from analyze import analyze

ROOT = Path(__file__).resolve().parent
KART_DIR = ROOT / "site_data" / "kart"
CARDREADS = ROOT / "cardreads"
TR_MON = ["", "Oca", "Şub", "Mar", "Nis", "May", "Haz", "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara"]
REPO = "hotelalsancak-netizen/hotelalsancak-netizen.github.io"


def log(msg):
    print(msg, flush=True)


def find_pdfs(root: Path):
    return sorted(Path(p) for p in glob.glob(str(root / "**" / "*.pdf"), recursive=True))


def decrypt_enc(src: Path) -> Path:
    """Decrypt a terminal-uploaded .zip.enc back to a plain .zip.

    The resepsiyon-PC uploader (terminal programı) encrypts the weekly zip before
    putting it in the public repo, so guest data is never public in the clear. Layout
    (must match the uploader): file = salt(16) + nonce(12) + ciphertext(+GCM tag);
    key = PBKDF2-HMAC-SHA256(CARD_ZIP_PASSWORD, salt, 200000). Same password on both
    machines (it lives only in .env, never in the repo)."""
    import hashlib
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    pw = os.environ.get("CARD_ZIP_PASSWORD", "").strip()
    if not pw:
        raise SystemExit("CARD_ZIP_PASSWORD .env'de yok. Terminal programındaki parolanın "
                         "AYNISINI .env'e ekleyin:  CARD_ZIP_PASSWORD=...")
    blob = src.read_bytes()
    salt, nonce, ct = blob[:16], blob[16:28], blob[28:]
    key = hashlib.pbkdf2_hmac("sha256", pw.encode(), salt, 200000, dklen=32)
    try:
        data = AESGCM(key).decrypt(nonce, ct, None)
    except Exception:
        raise SystemExit("Şifre çözülemedi — CARD_ZIP_PASSWORD yanlış olabilir "
                         "(terminal programındakiyle birebir aynı olmalı).")
    out = Path(tempfile.mkdtemp(prefix="kartdec_")) / src.name[:-4]  # <HAFTA>.zip.enc -> <HAFTA>.zip
    out.write_bytes(data)
    log(f"  🔓 .zip.enc çözüldü → {out.name} ({len(data):,} bayt)")
    return out


def extract_source(src: Path) -> Path:
    """Return a directory holding the PDFs (extracts a zip to a temp dir if needed).
    A terminal-uploaded .zip.enc is decrypted first (CARD_ZIP_PASSWORD)."""
    if src.is_dir():
        return src
    if src.suffix.lower() == ".enc":
        src = decrypt_enc(src)
    if src.suffix.lower() == ".zip":
        tmp = Path(tempfile.mkdtemp(prefix="kart_"))
        with zipfile.ZipFile(src) as z:
            z.extractall(tmp)
        log(f"  ZIP açıldı: {tmp}")
        return tmp
    raise SystemExit(f"Girdi bir .zip, .zip.enc veya klasör olmalı: {src}")


def parse_cards_folder(folder: Path) -> dict:
    pdfs = find_pdfs(folder)
    if not pdfs:
        raise SystemExit(f"PDF bulunamadı: {folder}")
    if not shutil.which("pdftotext"):
        raise SystemExit("pdftotext yok — 'brew install poppler' (kart PDF'leri için gerekli).")
    cards = {}
    for p in pdfs:
        room = p.stem.strip()
        try:
            cards[room] = parse_cards.parse_room(str(p))
        except Exception as e:
            log(f"  ! {p.name} okunamadı: {e}")
    log(f"  {len(pdfs)} PDF, {len(cards)} oda çözüldü")
    return cards


def _date_from_name(name: str):
    """Pull an export date out of the source name. Lock dumps hold ~2 months of
    history, so the week can NOT be inferred from the reads — it comes from when the
    export was taken. Supports 20.07.2026, 20260720(T…), 13072026."""
    import re
    for rx, order in ((r"(\d{2})\.(\d{2})\.(\d{4})", "dmy"),
                      (r"(\d{4})(\d{2})(\d{2})", "ymd"),
                      (r"(\d{2})(\d{2})(\d{4})", "dmy")):
        m = re.search(rx, name)
        if m:
            a, b, c = m.groups()
            try:
                return (dt.date(int(c), int(b), int(a)) if order == "dmy"
                        else dt.date(int(a), int(b), int(c)))
            except ValueError:
                continue
    return None


def week_of(override, src_name):
    """(monday, sunday) to publish. --week = any date IN the target week; otherwise
    take the export date from the file name and report the last FULLY-ENDED week —
    the most recent Mon–Sun period whose SUNDAY falls BEFORE the export date.

    Bug this fixes: the old code did `export - 1 day` and used that day's week, which
    only lands on the previous week when the export is on a MONDAY. Read the terminal
    any other day (e.g. Tue 04.08) and it picked the CURRENT, still-empty week
    (03–09) whose nights haven't happened yet → a meaningless "0 şüpheli". Now Tue
    04.08 correctly yields 27.07–02.08 (the week that is actually covered by data)."""
    if override:
        ref = dt.date.fromisoformat(override)
        monday = ref - dt.timedelta(days=ref.weekday())
        return monday, monday + dt.timedelta(days=6)
    exp = _date_from_name(src_name) or dt.date.today()
    last_sunday = exp - dt.timedelta(days=exp.weekday() + 1)   # most recent Sunday strictly before exp
    return last_sunday - dt.timedelta(days=6), last_sunday


def load_changes(src_dir: Path, lo: dt.date, hi: dt.date, extra_dirs=()) -> list:
    """Room changes for the week (best-effort). Looks for an "Oda Değişimi" export —
    Elektra's Excel/CSV or a JSON — inside the ZIP (or next to it), maps its Turkish
    headers, and filters to the week. Falls back to the repo's room_changes.json.
    Empty -> the report shows a warning banner (moved guests may read as suspicious)."""
    # 0) PRIMARY: pull the Oda Değişimi report straight from Elektra for this week.
    try:
        rows = E.fetch_room_changes(lo.isoformat(), hi.isoformat())
        log(f"  oda değişimi ELEKTRA'dan otomatik çekildi: {len(rows)} kayıt ({lo}–{hi})")
        return _filter_changes(rows, lo, hi)
    except Exception as e:
        log(f"  ! Elektra oda değişimi çekilemedi ({str(e)[:70]}) — dosya/yedeğe düşülüyor")

    exts = (".xlsx", ".xlsm", ".csv", ".json")
    kw = ("degis", "değiş", "roomchange", "room_change", "room-change", "odadeg", "oda deg")
    cands = []
    for p in Path(src_dir).rglob("*"):                    # zip içi: derin
        if p.is_file():
            cands.append(p)
    for d in extra_dirs:                                  # zip'in yanı: yüzeysel
        try:
            cands += [p for p in Path(d).glob("*") if p.is_file()]
        except (PermissionError, OSError):
            pass
    for f in cands:
        if f.suffix.lower() in exts and any(k in f.name.lower() for k in kw):
            try:
                rows = _read_changes(f)
            except SystemExit:
                raise
            except Exception as e:
                log(f"  ! {f.name} okunamadı: {e}"); continue
            if rows:
                wk = _filter_changes(rows, lo, hi)
                log(f"  oda değişimi dosyadan: {f.name} ({len(rows)} satır, {len(wk)} haftaya ait)")
                return wk
    rc = ROOT / "room_changes.json"                      # repo fallback
    if rc.exists():
        wk = _filter_changes([_canon_change(r) for r in json.loads(rc.read_text())], lo, hi)
        if wk:
            log(f"  oda değişimi room_changes.json'dan ({len(wk)} satır, haftaya filtreli)")
            return wk
    log("  ! oda değişimi verisi yok — rapora uyarı eklenecek (yanlış pozitif olabilir)")
    return []


def _canon_change(raw: dict) -> dict:
    """Map an arbitrary Oda-Değişimi row (Turkish/English headers) to the shape
    analyze.py wants: {when, guest, from_room, to_room, rez_id}."""
    import re
    out = {"when": "", "guest": "", "from_room": "", "to_room": "", "rez_id": ""}
    for k, v in raw.items():
        if k is None:
            continue
        h = re.sub(r"[^a-z0-9ğüşıöç ]", "", str(k).strip().lower())
        val = "" if v is None else str(v).strip()
        if not val:
            continue
        if not out["when"] and any(w in h for w in ("tarih", "saat", "zaman", "when", "date")):
            out["when"] = val
        elif "oda" in h or "room" in h:
            if any(w in h for w in ("eski", "onceki", "önceki", "cikis", "çıkış", "old", "from", "kaynak", "gelinen")):
                out["from_room"] = val
            elif any(w in h for w in ("yeni", "sonraki", "giris", "giriş", "new", "hedef", "gidilen")):
                out["to_room"] = val
        elif not out["guest"] and any(w in h for w in ("misafir", "musteri", "müşteri", "isim", "adsoyad", "ad soyad", "guest", "name")):
            out["guest"] = val
        elif not out["rez_id"] and any(w in h for w in ("rez", "reservation", "resid", "konaklama")):
            out["rez_id"] = val
    return out


def _read_changes(path: Path) -> list:
    """Read .xlsx/.csv/.json -> list of canonical change dicts."""
    suf = path.suffix.lower()
    if suf == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        raw = data if isinstance(data, list) else data.get("rows", [])
    elif suf in (".csv", ".txt"):
        import csv
        with path.open(encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(4096); fh.seek(0)
            delim = ";" if sample.count(";") > sample.count(",") else ","
            raw = [dict(r) for r in csv.DictReader(fh, delimiter=delim)]
    elif suf in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise SystemExit("Excel (Oda Değişimi) okumak için: pip install openpyxl")
        ws = load_workbook(path, read_only=True, data_only=True).active
        it = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else "" for c in next(it, [])]
        raw = [dict(zip(headers, r)) for r in it]
    else:
        return []
    return [_canon_change(r) for r in raw]


def _filter_changes(rows, lo, hi):
    out = []
    for r in rows:
        w = str(r.get("when", ""))[:10]
        try:
            d = dt.date.fromisoformat(w)
        except ValueError:
            out.append(r)   # keep undated rows rather than drop evidence
            continue
        if lo <= d <= hi:
            out.append(r)
    return out


def build_week(cards, changes, occ, lo, hi):
    # Attach each reservation's "Oda Notu" (ALLNOTES) to its room change, so the
    # report can show WHY the room change was done — an uncontrolled room change is
    # itself a leak risk, so the reason must be visible/auditable.
    try:
        notes = E.fetch_notes(sorted({c.get("rez_id") for c in changes if c.get("rez_id")}))
        for c in changes:
            c["note"] = notes.get(str(c.get("rez_id", "")), "")
        n = sum(1 for c in changes if c.get("note"))
        log(f"  oda notları: {n}/{len(changes)} değişimde neden yazılı")
    except Exception as ex:
        log(f"  ! oda notları çekilemedi ({str(ex)[:60]}) — nedensiz devam")
    html = build_report.build(cards, changes, occ, lo, hi)
    if not changes:
        banner = ('<div style="background:#fef3c7;border:1px solid #fde68a;color:#92400e;'
                  'padding:10px 14px;border-radius:10px;margin:0 0 14px;font-size:13px">'
                  '⚠️ Bu hafta için oda değişimi (Oda Değişimi raporu) yüklenmedi — '
                  'taşınan misafirler bazı odaları yanlışlıkla “şüpheli” gösterebilir.</div>')
        html = html.replace('<div class="wrap">', '<div class="wrap">' + banner, 1)
    findings = analyze(cards, occ, changes, lo.isoformat(), hi.isoformat())
    sus = sum(1 for f in findings if f["status"] == "SUSPICIOUS")
    return html, sus


def save_week(section, lo, hi, pw):
    KART_DIR.mkdir(parents=True, exist_ok=True)
    wid = lo.strftime("%Y%m%d")
    # Kart is MANAGER-ONLY: encrypt for the manager password only (multi-recipient
    # format so the role-aware shell reads it exactly like the cloud-built blobs).
    blob = dashcrypto.encrypt_multi(json.dumps(section, ensure_ascii=False), [pw])
    (KART_DIR / f"{wid}.enc.json").write_text(json.dumps(blob), encoding="utf-8")

    label = f"{lo.day}–{hi.day} {TR_MON[hi.month]} {hi.year}"
    idx_path = KART_DIR / "index.json"
    weeks = json.loads(idx_path.read_text())["weeks"] if idx_path.exists() else []
    weeks = [w for w in weeks if w["id"] != wid]
    weeks.append({"id": wid, "label": label, "file": f"kart-{wid}.enc.json",
                  "lo": lo.isoformat(), "hi": hi.isoformat()})
    weeks.sort(key=lambda w: w["id"], reverse=True)
    idx_path.write_text(json.dumps({"weeks": weeks}, ensure_ascii=False, indent=1), encoding="utf-8")
    return wid, label


def publish(wid):
    token = os.environ.get("GITHUB_TOKEN", "")
    env = dict(os.environ, GH_TOKEN=token)
    url = f"https://x-access-token:{token}@github.com/{REPO}.git"
    subprocess.run(["git", "-C", str(ROOT), "add", "site_data/kart"], check=True)
    r = subprocess.run(["git", "-C", str(ROOT), "commit", "-m",
                        f"Kart güvenliği: {wid} haftası"], env=env)
    if r.returncode != 0:
        log("  (commit edilecek değişiklik yok — yine de yayın tetikleniyor)")
    # Push alone triggers the workflow (on: push). No separate workflow_dispatch —
    # that double-fires and one run gets cancelled by the 'pages' concurrency group.
    subprocess.run(["git", "-C", str(ROOT), "push", url, "HEAD:main"], check=True, env=env)
    log("  push edildi → bulut yayını tetiklendi ✓")


def main():
    ap = argparse.ArgumentParser(description="Haftalık kart güvenliği yayını")
    ap.add_argument("source", help="Haftalık .zip veya PDF klasörü")
    ap.add_argument("--week", help="Hafta içindeki bir tarih YYYY-MM-DD (otomatik: son okumanın haftası)")
    ap.add_argument("--no-publish", action="store_true", help="Sadece yerelde üret, push etme")
    a = ap.parse_args()

    # Manager password (kart is manager-only). Mirrors dashboard.passwords():
    # DASH_PW_MANAGER, falling back to the legacy single DASH_PASSWORD.
    pw = (os.environ.get("DASH_PW_MANAGER") or os.environ.get("DASH_PASSWORD", "")).strip()
    if not pw:
        raise SystemExit("Yönetim şifresi yok (DASH_PW_MANAGER / DASH_PASSWORD, .env).")

    src = Path(os.path.expanduser(a.source))
    if not src.exists():
        raise SystemExit(f"Bulunamadı: {src}")

    log(f"1/6  Kaynak açılıyor: {src.name}")
    src_dir = extract_source(src)

    log("2/6  Kart PDF'leri çözülüyor...")
    cards = parse_cards_folder(src_dir)

    lo, hi = week_of(a.week, src.name)
    log(f"3/6  Hafta: {lo} – {hi}  (kaynak: {src.name})")

    # Kalıcı arşiv: cardreads/<YYYYMMDD>/
    dest = CARDREADS / lo.strftime("%Y%m%d")
    dest.mkdir(parents=True, exist_ok=True)
    for p in find_pdfs(src_dir):
        shutil.copy(p, dest / p.name)
    log(f"     PDF'ler arşivlendi: {dest.relative_to(ROOT)}")

    log("4/6  Oda planı (doluluk) Elektra'dan çekiliyor...")
    occ = E.fetch_room_calendar(lo.isoformat(), hi.isoformat())
    # Look inside the export AND next to it (an Oda Değişimi Excel dropped beside the zip).
    changes = load_changes(src_dir, lo, hi, extra_dirs=(src.parent,))

    log("5/6  Rapor üretiliyor + şifreleniyor...")
    html, sus = build_week(cards, changes, occ, lo, hi)
    section = {
        "label": "Haftalık Kart Güvenliği",
        "count": sus, "count_label": "şüpheli",
        "tone": "bad" if sus else "ok",
        "sub": f"{lo.day}–{hi.day} {TR_MON[hi.month]} {hi.year} · satılmadan kullanılan odalar",
        "updated": dt.datetime.now().strftime("%d.%m.%Y %H:%M"),
        "html": html,
    }
    wid, label = save_week(section, lo, hi, pw)
    log(f"     {label}: {sus} şüpheli oda-gecesi -> site_data/kart/{wid}.enc.json")

    if a.no_publish:
        log("6/6  --no-publish: yayın atlandı.")
    else:
        log("6/6  Yayınlanıyor...")
        publish(wid)
        log(f"Bitti. Birkaç dakikada dashboard'da '{label}' haftası görünür:")
        log("  https://hotelalsancak-netizen.github.io/  (Haftalık Kart Güvenliği)")


if __name__ == "__main__":
    sys.exit(main())
